from openpyxl import load_workbook

wb = load_workbook(filename = 'dados-2020-06-23.xlsx')

ws = wb["Plan1"]


#for row in ws:
#    for cell in row:
#        print(cell)
#    print(row)

sheet = wb.active
print(sheet.title)

row_count = sheet.max_row
column_count = sheet.max_column

print(column_count)

for i in range(2, row_count+1):
    str = sheet.cell(row=i, column=10).value
    if str[len(str)-4:len(str)] == '.pdf':
        sheet.cell(row=i, column=10).value = 'pdf_portal_ahm_atvdis_c02_ok.pdf'
        print('alterou ', str, ' para pdf_portal_ahm_atvdis_c02_ok.pdf')
    elif str[len(str)-4:len(str)] == '.mp3':
        sheet.cell(row=i, column=10).value = 'sample.mp3'
        print('alterou ', str, 'sample.mp3')
    elif str[len(str)-4:len(str)] == '.mp4':
        sheet.cell(row=i, column=10).value = 'm18_ARhis7_u01vt_lutero.mp4'
        print('alterou ', str, 'm18_ARhis7_u01vt_lutero.mp4')
        sheet.cell(row=i, column=36).value = 'm18_ARhis7_u01vt_lutero.vtt'
    else:
        sheet.cell(row=i, column=10).value = '/m18_ARmat6_u08jgm_gamamemoria'
        print('alterou ', str, '/m18_ARmat6_u08jgm_gamamemoria')


name = sheet.title+'.xlsx'
wb.save(name)